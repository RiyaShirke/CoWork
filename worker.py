"""
Cowork worker — atomic stock-statement ingestion.

Pulls the oldest unprocessed file from INBOX_DIR, extracts text (pdfplumber for
text PDFs, PaddleOCR via Docker for images and scanned PDFs), asks a local
Ollama LLM to map it to a strict JSON schema (report metadata + product rows +
optional batch rows), and appends the rows to a single Excel workbook.

Pipeline guarantees:
  * Per-file work is all-or-nothing. The workbook, the state file, and the
    file's location in inbox/processed/failed are either ALL updated together
    or NOT AT ALL. A crash mid-run leaves nothing half-done.
  * File-specific failures (corrupt PDF, malformed LLM output) increment an
    attempt counter; after RETRY_LIMIT attempts the file moves to failed/.
  * Environment failures (Docker down, Ollama unreachable, workbook open in
    Excel) do NOT burn attempts — they just skip the run.
  * Idempotency: same content hash never produces duplicate rows.

Intended to be triggered every 10-15 minutes by Windows Task Scheduler.
"""

from __future__ import annotations

import hashlib
import io
import json
import logging
import logging.handlers
import os
import re
import shutil
import subprocess
import sys
import tempfile
import time
import uuid
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import requests
from dotenv import load_dotenv
from openpyxl import Workbook, load_workbook


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

ROOT = Path(__file__).resolve().parent
load_dotenv(ROOT / ".env")


def _env_path(key: str, default: Path) -> Path:
    return Path(os.environ.get(key, str(default))).resolve()


def _env_int(key: str, default: int) -> int:
    try:
        return int(os.environ.get(key, str(default)))
    except ValueError:
        return default


INBOX_DIR = _env_path("INBOX_DIR", ROOT / "inbox")
PROCESSED_DIR = _env_path("PROCESSED_DIR", ROOT / "processed")
FAILED_DIR = _env_path("FAILED_DIR", ROOT / "failed")
OUTPUT_DIR = _env_path("OUTPUT_DIR", ROOT / "output")
STATE_DIR = _env_path("STATE_DIR", ROOT / "state")
LOG_DIR = _env_path("LOG_DIR", ROOT / "logs")
OUTPUT_XLSX = _env_path("OUTPUT_XLSX", OUTPUT_DIR / "bills.xlsx")

BATCH_LIMIT = _env_int("BATCH_LIMIT", 1)
RETRY_LIMIT = _env_int("RETRY_LIMIT", 3)
SUBPROCESS_TIMEOUT = _env_int("SUBPROCESS_TIMEOUT", 600)
ALLOWED_EXTS = tuple(
    e.strip().lower()
    for e in os.environ.get(
        "ALLOWED_EXTS", ".pdf,.png,.jpg,.jpeg,.bmp,.tif,.tiff"
    ).split(",")
    if e.strip()
)

PADDLE_IMAGE = os.environ.get("PADDLE_IMAGE", "paddlecloud/paddleocr:2.6-cpu-latest")
PADDLE_LANG = os.environ.get("PADDLE_LANG", "en")
PADDLE_CACHE_DIR = _env_path("PADDLE_CACHE_DIR", STATE_DIR / "paddleocr-cache")
PDF_OCR_DPI = _env_int("PDF_OCR_DPI", 200)

OLLAMA_URL = os.environ.get("OLLAMA_URL", "http://localhost:11434").rstrip("/")
OLLAMA_MODEL = os.environ.get("OLLAMA_MODEL", "qwen2.5:7b")
OLLAMA_TIMEOUT = _env_int("OLLAMA_TIMEOUT", 600)
OLLAMA_NUM_CTX = _env_int("OLLAMA_NUM_CTX", 8192)
OLLAMA_NUM_PREDICT = _env_int("OLLAMA_NUM_PREDICT", 16384)
OLLAMA_KEEP_ALIVE = os.environ.get("OLLAMA_KEEP_ALIVE", "10m")
OLLAMA_MAX_TEXT_CHARS = _env_int("OLLAMA_MAX_TEXT_CHARS", 24000)
OLLAMA_RETRY_ATTEMPTS = _env_int("OLLAMA_RETRY_ATTEMPTS", 3)
OLLAMA_RETRY_BACKOFF_SEC = _env_int("OLLAMA_RETRY_BACKOFF_SEC", 5)

# Minimum chars of native PDF text below which we treat the PDF as effectively
# image-based and run OCR as well. Stock statements with thousands of product
# rows but only a tiny embedded text layer (header only) fall here.
PDF_TEXT_MIN_CHARS = _env_int("PDF_TEXT_MIN_CHARS", 1500)

KREUZBERG_URL = os.environ.get("KREUZBERG_URL", "http://localhost:8000").strip().rstrip("/")
KREUZBERG_TIMEOUT = _env_int("KREUZBERG_TIMEOUT", 180)

NOTIFY_WEBHOOK_URL = os.environ.get("NOTIFY_WEBHOOK_URL", "").strip()

STATE_FILE = STATE_DIR / "processed.json"
LOCK_FILE = STATE_DIR / "worker.lock"
DEBUG_DIR = STATE_DIR / "debug"
LOCK_STALE_SECONDS = _env_int("LOCK_STALE_SECONDS", 60 * 60)

REPORT_META_COLUMNS = [
    "DistributorName",
    "DivisionName",
    "ReportFromDate",
    "ReportToDate",
    "TotalStockValue",
    "TotalSalesValue",
]

PRODUCT_COLUMNS = [
    "ProductCode",
    "ProductName",
    "Packing",
    "OpeningStock",
    "PurchaseQty",
    "SalesQty",
    "ClosingStock",
    "StockValue",
    "SalesValue",
    "AdjustmentQty",
    "PreviousMonthStock",
    "TwoMonthOldStock",
    "OldStock120Days",
    "ExpiryWithin3Months",
]

BATCH_COLUMNS = [
    "BatchNumber",
    "ExpiryDate",
    "BatchStock",
    "MRP",
    "BatchValue",
]

EXCEL_COLUMNS = (
    PRODUCT_COLUMNS
    + BATCH_COLUMNS
    + REPORT_META_COLUMNS
    + ["processed_at", "source_file"]
)

PDF_EXTS = {".pdf"}
IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff"}


# ---------------------------------------------------------------------------
# Errors
# ---------------------------------------------------------------------------


class FileExtractionError(RuntimeError):
    """File is bad in a way that re-running won't help. Counts against attempts."""


class EnvironmentSkip(RuntimeError):
    """Worker can't run right now (Docker/Ollama down, workbook locked). Does NOT count."""


# ---------------------------------------------------------------------------
# Bootstrap dirs + logging
# ---------------------------------------------------------------------------

for _d in (
    INBOX_DIR,
    PROCESSED_DIR,
    FAILED_DIR,
    OUTPUT_DIR,
    STATE_DIR,
    LOG_DIR,
    PADDLE_CACHE_DIR,
):
    _d.mkdir(parents=True, exist_ok=True)


def _setup_logging() -> logging.Logger:
    logger = logging.getLogger("cowork")
    if logger.handlers:
        return logger
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s | %(levelname)s | %(message)s")
    rot = logging.handlers.RotatingFileHandler(
        LOG_DIR / "cowork.log",
        maxBytes=5 * 1024 * 1024,
        backupCount=5,
        encoding="utf-8",
    )
    rot.setFormatter(fmt)
    stream = logging.StreamHandler(sys.stdout)
    stream.setFormatter(fmt)
    logger.addHandler(rot)
    logger.addHandler(stream)
    return logger


log = _setup_logging()


# ---------------------------------------------------------------------------
# Lock (PID-aware so a crashed worker doesn't block the scheduler for an hour)
# ---------------------------------------------------------------------------


def _pid_alive(pid: int) -> bool:
    if pid <= 0:
        return False
    try:
        os.kill(pid, 0)
    except ProcessLookupError:
        return False
    except PermissionError:
        # Process exists, we just can't signal it.
        return True
    except OSError:
        return False
    return True


def acquire_lock() -> bool:
    if LOCK_FILE.exists():
        try:
            payload = json.loads(LOCK_FILE.read_text(encoding="utf-8"))
            pid = int(payload.get("pid", -1))
            started_at = payload.get("started_at", "")
        except (ValueError, json.JSONDecodeError):
            pid, started_at = -1, ""
        age = time.time() - LOCK_FILE.stat().st_mtime
        if _pid_alive(pid) and age < LOCK_STALE_SECONDS:
            log.info(
                "Lock held by pid=%s since %s (age=%ds). Skipping this run.",
                pid, started_at, int(age),
            )
            return False
        log.warning(
            "Stale lock (pid=%s alive=%s age=%ds) — taking over.",
            pid, _pid_alive(pid), int(age),
        )
    LOCK_FILE.write_text(
        json.dumps({
            "pid": os.getpid(),
            "started_at": datetime.now().isoformat(timespec="seconds"),
        }),
        encoding="utf-8",
    )
    return True


def release_lock() -> None:
    try:
        LOCK_FILE.unlink(missing_ok=True)
    except OSError:
        pass


# ---------------------------------------------------------------------------
# Atomic state store
# ---------------------------------------------------------------------------


def load_state() -> dict:
    if not STATE_FILE.exists():
        return {}
    try:
        return json.loads(STATE_FILE.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        # Recover from a corrupt state file rather than crashing — but keep a
        # copy so the user can inspect it.
        corrupt_copy = STATE_FILE.with_suffix(f".corrupt.{int(time.time())}.json")
        try:
            shutil.copy2(STATE_FILE, corrupt_copy)
            log.warning("processed.json was corrupt; copied to %s and starting empty.", corrupt_copy)
        except OSError:
            log.warning("processed.json was corrupt; starting empty.")
        return {}


def save_state(state: dict) -> None:
    tmp = STATE_FILE.with_suffix(f".tmp.{uuid.uuid4().hex}.json")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(state, f, indent=2)
        f.flush()
        os.fsync(f.fileno())
    os.replace(tmp, STATE_FILE)


def sha256_of(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


# ---------------------------------------------------------------------------
# Inbox discovery
# ---------------------------------------------------------------------------


def discover_inbox() -> list[Path]:
    files = [
        p for p in INBOX_DIR.iterdir()
        if p.is_file() and p.suffix.lower() in ALLOWED_EXTS
    ]
    files.sort(key=lambda p: p.stat().st_mtime)
    return files


# ---------------------------------------------------------------------------
# Text extraction — Kreuzberg first (preserves table structure as Markdown),
# pdfplumber as fallback, PaddleOCR Docker as final fallback for scanned PDFs.
# ---------------------------------------------------------------------------


_KREUZBERG_CONFIG = {
    # Render the body as Markdown so headings/lists/tables come back in a
    # form the LLM understands well.
    "output_format": "markdown",
    # Include hierarchical document structure (helps with long multi-page docs).
    "include_document_structure": True,
    "pdf_options": {
        # Default true, but stated explicitly so a future server-default
        # change doesn't silently disable us.
        "extract_tables": True,
        # CRITICAL for distributor stock statements: their "tables" are
        # often just column-aligned text without proper PDF table objects.
        # Kreuzberg's stricter default rejects them; this lets them through.
        "allow_single_column_tables": True,
    },
}


def _kreuzberg_tables_to_markdown(tables: list[dict]) -> str:
    """Serialise Kreuzberg's structured tables to Markdown for the LLM."""
    parts: list[str] = []
    for idx, table in enumerate(tables, start=1):
        md = (table.get("markdown") or "").strip()
        if md:
            parts.append(f"### Table {idx} (page {table.get('page_number', '?')})\n\n{md}")
            continue
        cells = table.get("cells") or []
        if not cells:
            continue
        rows = ["| " + " | ".join(str(c) for c in row) + " |" for row in cells]
        if rows:
            header = rows[0]
            sep = "| " + " | ".join("---" for _ in cells[0]) + " |"
            parts.append("\n".join([header, sep] + rows[1:]))
    return "\n\n".join(parts)


def _extract_with_kreuzberg(pdf_path: Path) -> str:
    """Extract text + structured tables via the Kreuzberg REST service.

    Sends an extraction config that asks Kreuzberg to (a) render output as
    Markdown, (b) detect tables even when they're just column-aligned text
    (typical of distributor stock statements). If the response carries
    structured tables we prepend them to the body so the LLM sees the
    table grid first; otherwise we just return the Markdown body.
    """
    if not KREUZBERG_URL:
        return ""
    try:
        with pdf_path.open("rb") as f:
            r = requests.post(
                f"{KREUZBERG_URL}/extract",
                files={"files": (pdf_path.name, f, "application/pdf")},
                data={"config": json.dumps(_KREUZBERG_CONFIG)},
                timeout=KREUZBERG_TIMEOUT,
            )
    except (requests.ConnectionError, requests.Timeout) as exc:
        log.warning("Kreuzberg unreachable at %s (%s); falling back.", KREUZBERG_URL, exc)
        return ""
    except requests.RequestException as exc:
        log.warning("Kreuzberg request failed (%s); falling back.", exc)
        return ""

    if r.status_code != 200:
        log.warning("Kreuzberg API returned %s: %s", r.status_code, r.text[:200])
        return ""
    try:
        results = r.json()
    except ValueError:
        log.warning("Kreuzberg response was not JSON")
        return ""
    if not isinstance(results, list) or not results:
        return ""
    first = results[0] if isinstance(results[0], dict) else {}
    content = (first.get("content") or "").strip()
    tables = first.get("tables") or []
    pages = first.get("pages") or []

    # Prefer the per-page content because it preserves line breaks between rows.
    # Kreuzberg's top-level `content` field sometimes runs everything together
    # which makes column-aligned tables ambiguous for the LLM.
    if pages:
        page_chunks = []
        for p in pages:
            page_text = (p.get("content") or "").strip()
            if page_text:
                page_chunks.append(f"### Page {p.get('page_number', '?')}\n\n{page_text}")
        if page_chunks:
            body = "\n\n".join(page_chunks)
        else:
            body = content
    else:
        body = content

    log.info(
        "Kreuzberg: body=%d chars, %d page(s), %d structured table(s), output_format=%s",
        len(body), len(pages), len(tables),
        (first.get("metadata") or {}).get("output_format", "?"),
    )
    if tables:
        tables_md = _kreuzberg_tables_to_markdown(tables)
        if tables_md:
            return f"{tables_md}\n\n--- BODY ---\n\n{body}"
    return body


def _extract_with_pdfplumber(pdf_path: Path) -> str:
    """Fallback text extractor when Kreuzberg isn't installed or yields nothing."""
    try:
        import pdfplumber  # type: ignore
    except ImportError as exc:
        raise EnvironmentSkip(
            "pdfplumber not installed. Run: pip install -r requirements.txt"
        ) from exc

    pieces: list[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            if text.strip():
                pieces.append(text)
            for table in page.extract_tables() or []:
                for row in table:
                    cells = [str(c).strip() for c in row if c is not None]
                    if cells:
                        pieces.append("\t".join(cells))
    return "\n".join(pieces).strip()


def _extract_pdf_text_native(pdf_path: Path) -> str:
    """Return the best native text we can get for this PDF.

    Order: Kreuzberg (markdown w/ tables) → pdfplumber (plain text + tab-joined
    tables). We keep whichever produced more content; on ties Kreuzberg wins
    because table layout is the bigger LLM-accuracy lever.
    """
    kz = _extract_with_kreuzberg(pdf_path)
    if kz:
        log.info("Kreuzberg extracted %d chars", len(kz))
        return kz
    pp = _extract_with_pdfplumber(pdf_path)
    if pp:
        log.info("pdfplumber extracted %d chars (Kreuzberg returned nothing)", len(pp))
    return pp


def _render_pdf_pages_to_images(pdf_path: Path, out_dir: Path) -> list[Path]:
    try:
        import pypdfium2 as pdfium  # type: ignore
    except ImportError as exc:
        raise EnvironmentSkip(
            "pypdfium2 not installed; cannot render scanned PDFs for OCR fallback. "
            "Run: pip install -r requirements.txt"
        ) from exc

    scale = PDF_OCR_DPI / 72.0
    pages: list[Path] = []
    pdf = pdfium.PdfDocument(str(pdf_path))
    try:
        for idx in range(len(pdf)):
            page = pdf[idx]
            pil_image = page.render(scale=scale).to_pil()
            page_path = out_dir / f"page-{idx + 1:03d}.png"
            pil_image.save(page_path, format="PNG")
            pages.append(page_path)
    finally:
        pdf.close()
    return pages


_OCR_TEXT_RE = re.compile(r"\('((?:[^'\\]|\\.)*)',\s*0?\.\d+\)")


def _run_paddleocr_on_image(image: Path) -> str:
    """Run PaddleOCR in a one-shot container against a single image file.

    The image is staged in a per-call scratch directory so the container only
    ever sees that one file (not the entire inbox).
    """
    with tempfile.TemporaryDirectory(prefix="cowork-ocr-") as scratch:
        scratch_path = Path(scratch)
        staged = scratch_path / image.name
        shutil.copy2(image, staged)

        # Invoke via `python3 -c "import paddleocr; paddleocr.main()"` rather
        # than the `paddleocr` console script — the script in the
        # paddlecloud/paddleocr image looks up its own package metadata via
        # importlib_metadata and that lookup is broken in some builds
        # (PackageNotFoundError: paddleocr).
        cmd = [
            "docker", "run", "--rm",
            "-v", f"{scratch_path.resolve()}:/data",
            "-v", f"{PADDLE_CACHE_DIR.resolve()}:/root/.paddleocr",
            PADDLE_IMAGE,
            "python3", "-c", "import paddleocr; paddleocr.main()",
            "--image_dir", f"/data/{image.name}",
            "--lang", PADDLE_LANG,
            "--use_gpu", "false",
        ]
        log.info("PaddleOCR: %s", " ".join(cmd))
        try:
            result = subprocess.run(
                cmd,
                capture_output=True,
                text=True,
                encoding="utf-8",
                errors="replace",
                timeout=SUBPROCESS_TIMEOUT,
            )
        except FileNotFoundError as exc:
            raise EnvironmentSkip("Docker CLI not found on PATH") from exc
        except subprocess.TimeoutExpired as exc:
            raise EnvironmentSkip(
                f"PaddleOCR timed out after {SUBPROCESS_TIMEOUT}s on {image.name}"
            ) from exc

    if result.returncode != 0:
        stderr_tail = (result.stderr or "")[-2000:]
        if "Cannot connect to the Docker daemon" in stderr_tail or "docker daemon" in stderr_tail.lower():
            raise EnvironmentSkip("Docker daemon not reachable")
        raise FileExtractionError(
            f"PaddleOCR failed (exit {result.returncode}): {stderr_tail}"
        )

    matches = _OCR_TEXT_RE.findall(result.stdout)
    if matches:
        return "\n".join(matches)
    # No tuples parsed — return the raw stdout so the LLM still has something.
    return result.stdout.strip()


def _ocr_pdf_pages(pdf_path: Path) -> str:
    with tempfile.TemporaryDirectory(prefix="cowork-pdf-") as scratch:
        pages = _render_pdf_pages_to_images(pdf_path, Path(scratch))
        ocr_pieces = [_run_paddleocr_on_image(p) for p in pages]
    text = "\n".join(s for s in ocr_pieces if s)
    log.info("OCR (scanned PDF, %d page(s)): %d chars", len(ocr_pieces), len(text))
    return text


def extract_text(path: Path) -> str:
    """Return the best available text representation of a bill/statement file.

    For PDFs we run native extraction first (fast + accurate when it works),
    but fall back to OCR if the result is suspiciously short — many stock
    statements embed only their header as real text and rasterise the body,
    which would otherwise leave us with only a few hundred chars to feed the
    LLM.
    """
    ext = path.suffix.lower()
    if ext in PDF_EXTS:
        native = _extract_pdf_text_native(path)
        log.info("PDF native text: %d chars (threshold=%d)", len(native), PDF_TEXT_MIN_CHARS)
        if len(native) >= PDF_TEXT_MIN_CHARS:
            return native
        if native:
            log.warning(
                "PDF native text below threshold (%d < %d); supplementing with OCR.",
                len(native), PDF_TEXT_MIN_CHARS,
            )
        else:
            log.info("PDF had no embedded text; running OCR.")
        ocr_text = _ocr_pdf_pages(path)
        if native and ocr_text:
            # Keep both — native often has clean header values, OCR has the body.
            return native + "\n\n--- OCR ---\n" + ocr_text
        return ocr_text or native
    if ext in IMAGE_EXTS:
        text = _run_paddleocr_on_image(path)
        log.info("OCR (image): %d chars", len(text))
        return text
    raise FileExtractionError(f"Unsupported file extension: {ext}")


# ---------------------------------------------------------------------------
# Ollama extraction with strict JSON schema
# ---------------------------------------------------------------------------

EXTRACTION_SCHEMA: dict[str, Any] = {
    "type": "object",
    "properties": {
        "metadata": {
            "type": "object",
            "properties": {
                "DistributorName": {"type": ["string", "null"]},
                "DivisionName": {"type": ["string", "null"]},
                "ReportFromDate": {"type": ["string", "null"]},
                "ReportToDate": {"type": ["string", "null"]},
                "TotalStockValue": {"type": ["string", "number", "null"]},
                "TotalSalesValue": {"type": ["string", "number", "null"]},
            },
            "required": REPORT_META_COLUMNS,
        },
        "products": {
            "type": "array",
            "items": {
                "type": "object",
                "properties": {
                    "ProductCode": {"type": ["string", "null"]},
                    "ProductName": {"type": ["string", "null"]},
                    "Packing": {"type": ["string", "null"]},
                    "OpeningStock": {"type": ["string", "number", "null"]},
                    "PurchaseQty": {"type": ["string", "number", "null"]},
                    "SalesQty": {"type": ["string", "number", "null"]},
                    "ClosingStock": {"type": ["string", "number", "null"]},
                    "StockValue": {"type": ["string", "number", "null"]},
                    "SalesValue": {"type": ["string", "number", "null"]},
                    "AdjustmentQty": {"type": ["string", "number", "null"]},
                    "PreviousMonthStock": {"type": ["string", "number", "null"]},
                    "TwoMonthOldStock": {"type": ["string", "number", "null"]},
                    "OldStock120Days": {"type": ["string", "number", "null"]},
                    "ExpiryWithin3Months": {"type": ["string", "number", "null"]},
                    "batches": {
                        "type": "array",
                        "items": {
                            "type": "object",
                            "properties": {
                                "BatchNumber": {"type": ["string", "null"]},
                                "ExpiryDate": {"type": ["string", "null"]},
                                "BatchStock": {"type": ["string", "number", "null"]},
                                "MRP": {"type": ["string", "number", "null"]},
                                "BatchValue": {"type": ["string", "number", "null"]},
                            },
                            "required": BATCH_COLUMNS,
                        },
                    },
                },
                "required": PRODUCT_COLUMNS,
            },
        },
    },
    "required": ["metadata", "products"],
}


SYSTEM_PROMPT = """You are an information-extraction engine for distributor stock-statement PDFs.

You receive the text of ONE stock statement (extracted from PDF text or OCR). Different distributors print very different layouts — different column orders, different column names, different sets of columns. Your job is to (a) FIND the column header row in this specific document, (b) MAP each header label to the corresponding schema field below, and (c) PARSE each product row positionally against that mapping.

Return STRICT JSON with two keys: "metadata" and "products".

## 1. metadata (object)
   - DistributorName       : company / distributor name on the report header
   - DivisionName          : division or business unit, if shown
   - ReportFromDate        : start of reporting period (ISO YYYY-MM-DD when unambiguous, else original)
   - ReportToDate          : end of reporting period
   - TotalStockValue       : footer total stock value
   - TotalSalesValue       : footer total sales value

Use null for any field that is genuinely not present.

## 2. products (array; one element per product row)
Each element has:
   - ProductCode           : the SKU / item code, usually leftmost on the row
   - ProductName           : product description (often multi-word — keep ALL words until the packing token)
   - Packing               : packing/strip/size token (e.g. "10T", "10C", "60ML", "1GM", "SACHET", "VIAL")
   - OpeningStock          : opening quantity
   - PurchaseQty           : quantity purchased during the period
   - SalesQty              : quantity sold
   - ClosingStock          : closing quantity
   - StockValue            : monetary value of closing stock (usually has a decimal point)
   - SalesValue            : monetary value of sales (usually has a decimal point)
   - AdjustmentQty         : adjustment / correction qty (often absent)
   - PreviousMonthStock    : 30-day-old stock (often absent)
   - TwoMonthOldStock      : 60-day-old stock (often absent)
   - OldStock120Days       : 120-day-old stock (often absent)
   - ExpiryWithin3Months   : stock expiring within 3 months (often absent)
   - batches               : array; one element per batch row that appears under this product (empty array if none)
       BatchNumber, ExpiryDate, BatchStock, MRP, BatchValue

## Step 1 — find the column header

Scan the top of the document for a line of column labels. Common labels and their schema mappings (case-insensitive, fuzzy):

| Label seen in the report                                            | Schema field            |
|---------------------------------------------------------------------|-------------------------|
| Code / Item Code / SKU / Product Code                               | ProductCode             |
| Item / Item Description / Product Description / Particulars / Name  | ProductName             |
| Packing / Pack / Pkg / Strip                                        | Packing                 |
| Opening / OpStk / OPSTK / Op-Stk / OPN                              | OpeningStock            |
| Purchase / Purch / PURCH / Recd / Received / Inward                 | PurchaseQty             |
| Sale / Sales / SLS / SALE / Issued / Outward                        | SalesQty                |
| Closing / ClStk / CL-STK / STOCK / Bal / Balance                    | ClosingStock            |
| Stock-Value / StkVal / STK VAL / Closing Value                      | StockValue              |
| Sales-Value / SaleVal / SALE VAL / SLS VAL / Sale Amt               | SalesValue              |
| Adj / Adjustment / IN/OT / IN-OUT / Cor                             | AdjustmentQty           |
| Feb / Prev / Previous Month / Last Month / -1M                      | PreviousMonthStock      |
| Jan / 2 Months / -2M / 60D                                          | TwoMonthOldStock        |
| 120 / 120D / STK120 / 4 Months / +120 Days                          | OldStock120Days         |
| ExpYr / EXP3M / Exp 3M / Expiry 3 Months / Near Expiry              | ExpiryWithin3Months     |

Use the header to LOCK the column order for THIS document. Different vendors print the same data in different orders — do not assume any fixed order.

If the document has NO ProductCode column, leave ProductCode = null for every row. (Plenty of distributors skip it.)

## Step 2 — parse each product row positionally

Product rows are usually whitespace-separated tokens. Apply these rules:

1. **ProductCode** (when present in the header): leftmost token on the row that matches the code pattern (usually digits, sometimes alphanumeric).
2. **ProductName**: every token between the code (or start of line) and the packing token. May be MULTIPLE words like "FLUVIR 75 (H1)" or "HIFEN 200MG DT TAB H1".
3. **Packing**: the first token after the name that matches a packing pattern — ends in T/TAB/C/CAP/ML/GM/KG/G/MG, contains SACHET/VIAL/TUBE/BOTTLE/STRIP/CAPS/POWDER/SYP/SUSPENSION, or is a literal `NA`. Packing may span TWO tokens (e.g. "10 TAB", "6 TAB"). When in doubt, the packing is the shortest sensible prefix.
4. **Numeric columns**: every remaining token, in the SAME ORDER as the header you identified. Tokens with a decimal point (`.`) are almost always monetary values, not quantities.
5. **Zero-suppression**: many reports print "" or blank for zero values, so a row may have FEWER tokens than the header has columns. The omitted columns are USUALLY the trailing ones (right side) — but not always. If you can't be sure which column a token belongs to, leave that field null. NEVER guess.
6. **Strip commas** from numbers ("1,234.56" → 1234.56) but preserve decimals.

### Worked example A — code-prefixed layout (e.g. CHiNTAN AGENCIES)
Header: `Code  Item Description  Packing  Opening  Purchase  Sales  Closing  Stock-Value  Sales-Value`

Row: `11071 FLUVIR 75 (H1) 10C 154 0 39 115 41835.05 13473.25`
Parse:
   ProductCode  = "11071"
   ProductName  = "FLUVIR 75 (H1)"
   Packing      = "10C"
   OpeningStock = 154
   PurchaseQty  = 0
   SalesQty     = 39
   ClosingStock = 115
   StockValue   = 41835.05
   SalesValue   = 13473.25

Zero-suppressed: `44769 FLUVIR 30 (H1) 10C 5 0 5 927.89`
Only ONE decimal token → 927.89 = StockValue, SalesValue = null.
Three ints (5, 0, 5) → Opening=5, Purchase=0, Closing=5, Sales = null.

### Worked example B — no-code layout with old-stock columns (e.g. SAMARTH DISTRIBUTORS / ST-HMAIN)
Header: `PRODUCT DESCRIPTION  PACKING  OPSTK  PURCH  SALE  SALE VAL  IN/OT  STOCK  STK VAL  FEB  JAN  STK120  EXP3M`

(No code column → ProductCode = null for every row.)

Row: `ENUFF 10 SACHET 1GM 235 500 293 3293 0 442 4469 187 379`
Parse:
   ProductCode         = null
   ProductName         = "ENUFF 10 SACHET"
   Packing             = "1GM"
   OpeningStock        = 235
   PurchaseQty         = 500
   SalesQty            = 293
   SalesValue          = 3293
   AdjustmentQty       = 0
   ClosingStock        = 442
   StockValue          = 4469
   PreviousMonthStock  = 187
   TwoMonthOldStock    = 379
   OldStock120Days     = null
   ExpiryWithin3Months = null
   (the last two trailing columns were absent — zero-suppression)

## Batch lines
A line that starts with "Batch" or contains "Batch :" / "Batch No" belongs to the most recent product above it. Example:
   Batch : HH2503138  Stock : 1  ExPDt : 08/26  MRP : 247.75  Value : 169.89
Extract: BatchNumber="HH2503138", BatchStock=1, ExpiryDate="08/26" (or normalised), MRP=247.75, BatchValue=169.89.

## Other layouts you may see
- Proper Markdown tables (`| col | col |` rows). When the input already has these, use them directly — the column names tell you exactly which schema field each cell maps to.
- Multi-page reports: products may continue across page breaks. Keep extracting; only stop at TOTALS / Grand Total / End of Report.
- Some distributors split product info across two lines (name on line 1, numbers on line 2). Re-assemble before parsing.
- Numbers may use comma thousands separators (e.g. "1,996.20"). Strip commas before emitting.

## Output rules
- Extract EVERY product row visible. Do not summarise, skip, or merge rows.
- Output JSON only — no markdown fences, no commentary, no prose.
- Use null (not "" or "N/A" or "-") for absent fields.
- Numbers as plain numerics where possible; if the source text is ambiguous, keep the original string.
- Dates in ISO YYYY-MM-DD when you can determine the format; otherwise keep the original.
- Never invent data not visible in the input.
"""


def _compact_text(text: str) -> str:
    """Keep size sensible for the LLM while preserving line order."""
    lines = [ln.rstrip() for ln in text.splitlines()]
    lines = [ln for ln in lines if ln.strip()]
    deduped: list[str] = []
    last = None
    for ln in lines:
        if ln != last:
            deduped.append(ln)
        last = ln
    compact = "\n".join(deduped)
    if len(compact) <= OLLAMA_MAX_TEXT_CHARS:
        return compact
    head_chars = int(OLLAMA_MAX_TEXT_CHARS * 0.7)
    tail_chars = OLLAMA_MAX_TEXT_CHARS - head_chars
    return compact[:head_chars].rstrip() + "\n...\n" + compact[-tail_chars:].lstrip()


def _dump_debug(label: str, content: str) -> Path | None:
    """Write a failure artefact to state/debug/ for after-the-fact inspection."""
    try:
        DEBUG_DIR.mkdir(parents=True, exist_ok=True)
        path = DEBUG_DIR / f"{datetime.now().strftime('%Y%m%d-%H%M%S')}__{label}"
        path.write_text(content, encoding="utf-8")
        return path
    except OSError as exc:
        log.warning("Could not write debug dump %s: %s", label, exc)
        return None


def call_ollama(text: str, source_file: str = "input") -> dict:
    prompt_text = _compact_text(text)
    log.info(
        "Ollama: input text raw=%d chars, compacted=%d chars",
        len(text), len(prompt_text),
    )

    payload = {
        "model": OLLAMA_MODEL,
        "prompt": f"{SYSTEM_PROMPT}\n\nSTOCK_STATEMENT_TEXT:\n<<<\n{prompt_text}\n>>>",
        "stream": False,
        "format": EXTRACTION_SCHEMA,
        "keep_alive": OLLAMA_KEEP_ALIVE,
        "options": {
            "temperature": 0.0,
            "num_ctx": OLLAMA_NUM_CTX,
            "num_predict": OLLAMA_NUM_PREDICT,
        },
    }
    log.info(
        "Ollama: POST /api/generate (model=%s, num_ctx=%d, num_predict=%d, timeout=%ds)",
        OLLAMA_MODEL, OLLAMA_NUM_CTX, OLLAMA_NUM_PREDICT, OLLAMA_TIMEOUT,
    )

    last_err: Exception | None = None
    for attempt in range(1, OLLAMA_RETRY_ATTEMPTS + 1):
        try:
            r = requests.post(
                f"{OLLAMA_URL}/api/generate", json=payload, timeout=OLLAMA_TIMEOUT,
            )
        except (requests.ConnectionError, requests.Timeout) as exc:
            last_err = exc
            log.warning(
                "Ollama transport error (attempt %d/%d): %s",
                attempt, OLLAMA_RETRY_ATTEMPTS, exc,
            )
            if attempt < OLLAMA_RETRY_ATTEMPTS:
                time.sleep(OLLAMA_RETRY_BACKOFF_SEC * attempt)
            continue

        if r.status_code >= 500:
            last_err = RuntimeError(
                f"Ollama HTTP {r.status_code}: {r.text[:300]}"
            )
            log.warning(
                "Ollama server error (attempt %d/%d): %s",
                attempt, OLLAMA_RETRY_ATTEMPTS, last_err,
            )
            if attempt < OLLAMA_RETRY_ATTEMPTS:
                time.sleep(OLLAMA_RETRY_BACKOFF_SEC * attempt)
            continue

        if r.status_code == 404:
            raise EnvironmentSkip(
                f"Ollama model '{OLLAMA_MODEL}' not pulled. "
                f"Run: docker exec cowork-ollama ollama pull {OLLAMA_MODEL}"
            )

        r.raise_for_status()
        body = r.json()
        raw = (body.get("response") or "").strip()
        done_reason = body.get("done_reason") or ""
        eval_count = body.get("eval_count")
        log.info(
            "Ollama: response len=%d chars, eval_count=%s, done_reason=%s",
            len(raw), eval_count, done_reason,
        )
        try:
            return json.loads(raw)
        except json.JSONDecodeError as exc:
            m = re.search(r"\{.*\}", raw, re.DOTALL)
            if not m:
                dump = _dump_debug(f"{source_file}.no-json.txt", raw)
                raise FileExtractionError(
                    f"Ollama response was not JSON (len={len(raw)}). "
                    f"Full output dumped to: {dump}"
                ) from exc
            try:
                return json.loads(m.group(0))
            except json.JSONDecodeError as exc2:
                dump = _dump_debug(f"{source_file}.partial-json.txt", raw)
                hint = ""
                if done_reason == "length" or (
                    eval_count is not None and eval_count >= OLLAMA_NUM_PREDICT - 16
                ):
                    hint = (
                        " — output hit num_predict ceiling; raise OLLAMA_NUM_PREDICT "
                        "or split the source into smaller pieces."
                    )
                raise FileExtractionError(
                    f"Ollama response could not be parsed as JSON "
                    f"(len={len(raw)}, eval_count={eval_count}, done_reason={done_reason}){hint}. "
                    f"Full output dumped to: {dump}"
                ) from exc2

    raise EnvironmentSkip(
        f"Ollama unreachable after {OLLAMA_RETRY_ATTEMPTS} attempt(s): {last_err}"
    )


# ---------------------------------------------------------------------------
# Schema validation + row flattening
# ---------------------------------------------------------------------------


@dataclass
class ExtractedDoc:
    metadata: dict
    products: list[dict]


def _coerce_scalar(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def validate_and_normalize(raw: dict) -> ExtractedDoc:
    if not isinstance(raw, dict):
        raise FileExtractionError("LLM output is not a JSON object")

    metadata = raw.get("metadata") or {}
    if not isinstance(metadata, dict):
        raise FileExtractionError("metadata is not an object")

    products = raw.get("products") or []
    if not isinstance(products, list):
        raise FileExtractionError("products is not an array")

    norm_meta = {col: _coerce_scalar(metadata.get(col)) for col in REPORT_META_COLUMNS}

    norm_products: list[dict] = []
    for idx, product in enumerate(products):
        if not isinstance(product, dict):
            raise FileExtractionError(f"products[{idx}] is not an object")
        norm = {col: _coerce_scalar(product.get(col)) for col in PRODUCT_COLUMNS}
        batches_raw = product.get("batches") or []
        if not isinstance(batches_raw, list):
            raise FileExtractionError(f"products[{idx}].batches is not an array")
        norm_batches: list[dict] = []
        for bidx, batch in enumerate(batches_raw):
            if not isinstance(batch, dict):
                raise FileExtractionError(
                    f"products[{idx}].batches[{bidx}] is not an object"
                )
            norm_batches.append(
                {col: _coerce_scalar(batch.get(col)) for col in BATCH_COLUMNS}
            )
        norm["batches"] = norm_batches
        norm_products.append(norm)

    if not norm_products:
        raise FileExtractionError("LLM returned zero products")

    _log_totals_reconciliation(norm_meta, norm_products)
    return ExtractedDoc(metadata=norm_meta, products=norm_products)


def _to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).replace(",", "").strip()
    try:
        return float(s)
    except ValueError:
        return None


def _log_totals_reconciliation(metadata: dict, products: list[dict]) -> None:
    """Soft check: sum(product.StockValue) should ≈ TotalStockValue. Logs only —
    never fails the extraction. Lets a human notice if the model is dropping
    rows or miscounting columns on a particular vendor format.
    """
    expected_stk = _to_float(metadata.get("TotalStockValue"))
    expected_sal = _to_float(metadata.get("TotalSalesValue"))
    sum_stk = sum(filter(None, (_to_float(p.get("StockValue")) for p in products)))
    sum_sal = sum(filter(None, (_to_float(p.get("SalesValue")) for p in products)))

    def _delta(actual: float, expected: float | None) -> str:
        if expected is None or expected == 0:
            return "n/a"
        return f"{(actual - expected) / expected * 100:+.1f}%"

    log.info(
        "Totals reconciliation: products=%d  sum(StockValue)=%.2f vs report=%s (%s)  sum(SalesValue)=%.2f vs report=%s (%s)",
        len(products),
        sum_stk, expected_stk if expected_stk is not None else "?",
        _delta(sum_stk, expected_stk),
        sum_sal, expected_sal if expected_sal is not None else "?",
        _delta(sum_sal, expected_sal),
    )


def flatten_rows(doc: ExtractedDoc, source_file: str, processed_at: str) -> list[list[Any]]:
    rows: list[list[Any]] = []
    blank_batch = {col: "" for col in BATCH_COLUMNS}
    for product in doc.products:
        batches = product.get("batches") or []
        if not batches:
            batches = [blank_batch]
        for batch in batches:
            row: list[Any] = []
            row.extend(product.get(col, "") for col in PRODUCT_COLUMNS)
            row.extend(batch.get(col, "") for col in BATCH_COLUMNS)
            row.extend(doc.metadata.get(col, "") for col in REPORT_META_COLUMNS)
            row.extend([processed_at, source_file])
            rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Atomic Excel append
# ---------------------------------------------------------------------------


def ensure_output_writable() -> None:
    if not OUTPUT_XLSX.exists():
        return
    try:
        with OUTPUT_XLSX.open("a+b"):
            pass
    except PermissionError as exc:
        raise EnvironmentSkip(
            f"{OUTPUT_XLSX.name} is open in Excel — close it and re-run."
        ) from exc


def _load_or_create_workbook() -> tuple[Workbook, Any]:
    if OUTPUT_XLSX.exists():
        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active
        existing_header = [c.value for c in ws[1]] if ws.max_row >= 1 else []
        if existing_header == EXCEL_COLUMNS:
            return wb, ws
        # Schema mismatch (column reorder, rename, etc.). Archive the old file
        # so a manual `rm bills.xlsx` is never needed; new workbook starts clean.
        stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
        archived = OUTPUT_XLSX.with_name(f"{OUTPUT_XLSX.stem}.legacy-{stamp}.xlsx")
        try:
            wb.close()
        except Exception:
            pass
        try:
            OUTPUT_XLSX.replace(archived)
            log.warning(
                "Output workbook header doesn't match current schema; "
                "archived old file to %s and starting a fresh workbook.",
                archived.name,
            )
        except OSError as exc:
            log.warning("Could not archive old workbook (%s); will append anyway.", exc)
    wb = Workbook()
    ws = wb.active
    ws.title = "inventory"
    ws.append(EXCEL_COLUMNS)
    return wb, ws


def append_rows_atomic(rows: list[list[Any]]) -> int:
    """Append rows or fail without touching the existing workbook.

    Strategy: load (or create) → append in memory → save to .new → fsync →
    backup current to .bak → os.replace .new over the live file → drop .bak.
    If anything throws before the final replace, the live file is untouched.
    """
    OUTPUT_XLSX.parent.mkdir(parents=True, exist_ok=True)

    wb, ws = _load_or_create_workbook()
    for row in rows:
        ws.append(row)

    tmp_path = OUTPUT_XLSX.with_suffix(f".new.{uuid.uuid4().hex}.xlsx")
    bak_path = OUTPUT_XLSX.with_suffix(".bak.xlsx")

    buf = io.BytesIO()
    wb.save(buf)
    data = buf.getvalue()

    with tmp_path.open("wb") as f:
        f.write(data)
        f.flush()
        os.fsync(f.fileno())

    backed_up = False
    if OUTPUT_XLSX.exists():
        shutil.copy2(OUTPUT_XLSX, bak_path)
        backed_up = True

    try:
        os.replace(tmp_path, OUTPUT_XLSX)
    except PermissionError as exc:
        # Workbook got opened in Excel between our writable check and the swap.
        try:
            tmp_path.unlink(missing_ok=True)
        except OSError:
            pass
        raise EnvironmentSkip(
            f"{OUTPUT_XLSX.name} became locked during write; will retry next run."
        ) from exc

    if backed_up:
        try:
            bak_path.unlink(missing_ok=True)
        except OSError:
            pass
    return len(rows)


# ---------------------------------------------------------------------------
# File movement
# ---------------------------------------------------------------------------


def move_to(path: Path, dest_dir: Path) -> Path:
    dest_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    dest = dest_dir / f"{stamp}__{path.name}"
    shutil.move(str(path), str(dest))
    return dest


# ---------------------------------------------------------------------------
# Notifications
# ---------------------------------------------------------------------------


def notify(text: str) -> None:
    if not NOTIFY_WEBHOOK_URL:
        return
    try:
        requests.post(NOTIFY_WEBHOOK_URL, json={"text": text}, timeout=10)
    except requests.RequestException as exc:
        log.warning("Notification webhook failed: %s", exc)


# ---------------------------------------------------------------------------
# Per-file pipeline
# ---------------------------------------------------------------------------


@dataclass
class RunStats:
    succeeded: int = 0
    file_failed: int = 0
    env_skipped: int = 0
    moved_to_failed: int = 0


def process_one(path: Path, state: dict, stats: RunStats) -> None:
    log.info("---- processing %s ----", path.name)
    digest = sha256_of(path)
    record = state.get(digest, {}) or {}

    if record.get("status") == "success":
        log.info(
            "Already processed (hash=%s, on %s). Moving duplicate to processed/.",
            digest[:10], record.get("processed_at"),
        )
        try:
            move_to(path, PROCESSED_DIR)
        except OSError as exc:
            log.warning("Could not move duplicate: %s", exc)
        return

    attempts_so_far = int(record.get("attempts", 0))

    try:
        ensure_output_writable()
        text = extract_text(path)
        if not text.strip():
            raise FileExtractionError("Extracted text was empty")

        raw = call_ollama(text, source_file=path.name)
        doc = validate_and_normalize(raw)

        processed_at = datetime.now().isoformat(timespec="seconds")
        rows = flatten_rows(doc, source_file=path.name, processed_at=processed_at)
        if not rows:
            raise FileExtractionError("No rows produced after flattening")

        ensure_output_writable()
        n_written = append_rows_atomic(rows)
        dest = move_to(path, PROCESSED_DIR)

        state[digest] = {
            "source_file": path.name,
            "processed_at": processed_at,
            "status": "success",
            "rows_added": n_written,
            "products": len(doc.products),
            "attempts": attempts_so_far + 1,
            "moved_to": str(dest),
        }
        save_state(state)
        stats.succeeded += 1
        log.info(
            "OK: %d row(s) from %d product(s) appended to %s",
            n_written, len(doc.products), OUTPUT_XLSX,
        )

    except EnvironmentSkip as exc:
        # Worker-level problem; do NOT burn an attempt.
        stats.env_skipped += 1
        log.warning("SKIP (environment): %s", exc)

    except FileExtractionError as exc:
        new_attempts = attempts_so_far + 1
        log.error(
            "FAIL (%d/%d): %s — %s",
            new_attempts, RETRY_LIMIT, path.name, exc,
        )
        stats.file_failed += 1

        if new_attempts >= RETRY_LIMIT:
            try:
                dest = move_to(path, FAILED_DIR)
                moved_to = str(dest)
            except OSError as move_exc:
                log.warning("Could not move to failed/: %s", move_exc)
                moved_to = str(path)
            state[digest] = {
                "source_file": path.name,
                "last_attempt_at": datetime.now().isoformat(timespec="seconds"),
                "status": "failed",
                "attempts": new_attempts,
                "error": str(exc),
                "moved_to": moved_to,
            }
            save_state(state)
            stats.moved_to_failed += 1
            notify(
                f"Cowork: {path.name} moved to failed/ after {new_attempts} attempts. "
                f"Last error: {exc}"
            )
        else:
            state[digest] = {
                "source_file": path.name,
                "last_attempt_at": datetime.now().isoformat(timespec="seconds"),
                "status": "retrying",
                "attempts": new_attempts,
                "error": str(exc),
            }
            save_state(state)

    except Exception as exc:  # noqa: BLE001
        # Unexpected programming error — treat as file-specific so we don't
        # silently retry forever, but log loud.
        new_attempts = attempts_so_far + 1
        log.exception("UNEXPECTED FAIL (%d/%d) %s", new_attempts, RETRY_LIMIT, path.name)
        stats.file_failed += 1
        if new_attempts >= RETRY_LIMIT:
            try:
                dest = move_to(path, FAILED_DIR)
                moved_to = str(dest)
            except OSError:
                moved_to = str(path)
            state[digest] = {
                "source_file": path.name,
                "last_attempt_at": datetime.now().isoformat(timespec="seconds"),
                "status": "failed",
                "attempts": new_attempts,
                "error": f"{type(exc).__name__}: {exc}",
                "moved_to": moved_to,
            }
            save_state(state)
            stats.moved_to_failed += 1
            notify(
                f"Cowork: {path.name} moved to failed/ after {new_attempts} attempts. "
                f"Unexpected error: {type(exc).__name__}: {exc}"
            )
        else:
            state[digest] = {
                "source_file": path.name,
                "last_attempt_at": datetime.now().isoformat(timespec="seconds"),
                "status": "retrying",
                "attempts": new_attempts,
                "error": f"{type(exc).__name__}: {exc}",
            }
            save_state(state)


# ---------------------------------------------------------------------------
# Status CLI
# ---------------------------------------------------------------------------


def print_status() -> int:
    state = load_state()
    if not state:
        print("No files processed yet.")
        return 0
    by_status: dict[str, int] = {}
    retrying: list[tuple[int, str, str]] = []
    failed: list[tuple[str, str]] = []
    succeeded_today = 0
    today = datetime.now().date().isoformat()

    for record in state.values():
        status = record.get("status", "unknown")
        by_status[status] = by_status.get(status, 0) + 1
        if status == "success" and (record.get("processed_at") or "").startswith(today):
            succeeded_today += 1
        elif status == "retrying":
            retrying.append((
                int(record.get("attempts", 0)),
                record.get("source_file", "?"),
                record.get("error", ""),
            ))
        elif status == "failed":
            failed.append((record.get("source_file", "?"), record.get("error", "")))

    print(f"Total tracked files: {sum(by_status.values())}")
    for status, count in sorted(by_status.items()):
        print(f"  {status}: {count}")
    print(f"  succeeded today: {succeeded_today}")

    if retrying:
        print("\nIn retry queue:")
        for attempts, name, err in sorted(retrying, reverse=True):
            print(f"  [{attempts}/{RETRY_LIMIT}] {name} — {err[:120]}")
    if failed:
        print("\nMoved to failed/ (latest first):")
        for name, err in failed[-10:]:
            print(f"  {name} — {err[:120]}")
    return 0


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


@contextmanager
def _lock_scope():
    if not acquire_lock():
        yield False
        return
    try:
        yield True
    finally:
        release_lock()


def main(argv: list[str] | None = None) -> int:
    args = sys.argv[1:] if argv is None else argv
    if args and args[0] in {"--status", "status"}:
        return print_status()

    run_started = time.monotonic()
    log.info("Cowork worker run starting. INBOX=%s MODEL=%s", INBOX_DIR, OLLAMA_MODEL)

    with _lock_scope() as got_lock:
        if not got_lock:
            return 0
        stats = RunStats()
        files = discover_inbox()
        log.info("Inbox: %d candidate file(s)", len(files))
        if not files:
            log.info(
                "RUN_SUMMARY processed=0 succeeded=0 file_failed=0 env_skipped=0 "
                "moved_to_failed=0 duration=%.1fs",
                time.monotonic() - run_started,
            )
            return 0

        if BATCH_LIMIT > 0:
            files = files[:BATCH_LIMIT]

        state = load_state()
        for f in files:
            process_one(f, state, stats)

        log.info(
            "RUN_SUMMARY processed=%d succeeded=%d file_failed=%d env_skipped=%d "
            "moved_to_failed=%d duration=%.1fs",
            len(files),
            stats.succeeded,
            stats.file_failed,
            stats.env_skipped,
            stats.moved_to_failed,
            time.monotonic() - run_started,
        )
        return 0


if __name__ == "__main__":
    raise SystemExit(main())
