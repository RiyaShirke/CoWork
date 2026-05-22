from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

import sys

ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(ROOT))

import worker  # noqa: E402


SOURCE_FILE = "WhatsApp Image 2026-05-21 at 11.19.24.jpeg"
PROCESSED_FILE = ROOT / "processed" / f"20260521-115251__{SOURCE_FILE}"


def main() -> int:
    if not PROCESSED_FILE.exists():
        print(f"Missing processed invoice: {PROCESSED_FILE}")
        return 1

    try:
        worker.ensure_output_writable()
    except worker.OutputWorkbookLocked as exc:
        print(exc)
        return 1
    ocr_text = worker.run_paddleocr(PROCESSED_FILE)
    rows = worker.fallback_extract_rows(ocr_text)

    wb = load_workbook(worker.OUTPUT_XLSX)
    ws = wb.active
    source_col = worker.EXCEL_COLUMNS.index("source_file") + 1

    for row_idx in range(ws.max_row, 1, -1):
        if ws.cell(row_idx, source_col).value == SOURCE_FILE:
            ws.delete_rows(row_idx, 1)

    now_iso = datetime.now().isoformat(timespec="seconds")
    for row in rows:
        ws.append([
            now_iso,
            SOURCE_FILE,
            row.get("vendor", "") or "",
            row.get("invoice_no", "") or "",
            row.get("invoice_date", "") or "",
            row.get("line_item", "") or "",
            row.get("quantity", "") if row.get("quantity") is not None else "",
            row.get("unit_price", "") if row.get("unit_price") is not None else "",
            row.get("amount", "") if row.get("amount") is not None else "",
            row.get("currency", "") or "",
            row.get("notes", "") or "",
        ])

    wb.save(worker.OUTPUT_XLSX)
    print(f"Replaced Fresh Express rows with {len(rows)} corrected row(s).")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
